from openpyxl import load_workbook
from column_header_men import create_column_header_men
from column_header_women import create_column_header_women
import os


def write_spreadsheet(content, row_num, gender):  # content is a list
    if gender == "M":
        file_path = './data/men.xlsx'
        # Check if the path exists
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
        else:
            # create column header
            file_path = create_column_header_men()
            # load workbook
            workbook = load_workbook(file_path)

        # Access the active sheet
        sheet = workbook.active

        # Sample data for a row
        row_data = content

        # Write data horizontally
        row_num = row_num
        for col_num, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value

        # Save the modified workbook
        workbook.save('./data/men.xlsx')

    elif gender == "W":
        file_path = './data/women.xlsx'
        # Check if the path exists
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
        else:
            # create column header
            file_path = create_column_header_women()
            # load workbook
            workbook = load_workbook(file_path)

        # Access the active sheet
        sheet = workbook.active

        # Sample data for a row
        row_data = content

        # Write data horizontally
        row_num = row_num
        for col_num, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value

        # Save the modified workbook
        workbook.save('./data/women.xlsx')



