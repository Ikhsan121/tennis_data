from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import os


def create_column_header_women():
    # Specify the directory path
    directory = './data'

    # Create the directory if it doesn't exist
    if not os.path.exists(directory):
        os.makedirs(directory)
    # Create a new workbook
    workbook = Workbook()
    sheet = workbook.active

    # Set player field
    sheet['A2'] = 'Player'

    ##############################################
    # Merge cells for Career-Tour
    sheet.merge_cells('B1:E1')
    sheet['B1'] = 'Career-Tour'
    sheet['B1'].alignment = Alignment(horizontal='center')

    # Set column headers and values for Career-Tour
    headers = ['M', 'Win%', 'MS', 'TPW']
    for col_num, header in enumerate(headers, start=2):
        col_letter = get_column_letter(col_num)
        sheet[col_letter + '2'] = header
    #############################################
    # Merge cells for Tour-2023
    sheet.merge_cells('F1:I1')
    sheet['F1'] = 'Tour-2023'
    sheet['F1'].alignment = Alignment(horizontal='center')

    # Set column headers and values for Tour-2023
    for col_num, header in enumerate(headers, start=6):
        col_letter = get_column_letter(col_num)
        sheet[col_letter + '2'] = header
    ###########################################
    # Merge cells for Tour-2022
    sheet.merge_cells('J1:M1')
    sheet['J1'] = 'Tour-2022'
    sheet['J1'].alignment = Alignment(horizontal='center')

    # Set column headers and values for Tour-2022
    for col_num, header in enumerate(headers, start=10):
        col_letter = get_column_letter(col_num)
        sheet[col_letter + '2'] = header
    ##########################################
    # Merge cells for Career-Hard
    sheet.merge_cells('N1:Q1')
    sheet['N1'] = 'Career-Hard'
    sheet['N1'].alignment = Alignment(horizontal='center')

    # Set column headers and values for Career-Hard
    for col_num, header in enumerate(headers, start=14):
        col_letter = get_column_letter(col_num)
        sheet[col_letter + '2'] = header
    ###########################################
    # Merge cells for Career-Clay
    sheet.merge_cells('R1:U1')
    sheet['R1'] = 'Career-Clay'
    sheet['R1'].alignment = Alignment(horizontal='center')

    # Set column headers and values for Career-Clay
    for col_num, header in enumerate(headers, start=18):
        col_letter = get_column_letter(col_num)
        sheet[col_letter + '2'] = header
    ##########################################
    # Merge cells for Career-Grass
    sheet.merge_cells('V1:Y1')
    sheet['V1'] = 'Career-Grass'
    sheet['V1'].alignment = Alignment(horizontal='center')

    # Set column headers and values for Career-Tour
    for col_num, header in enumerate(headers, start=22):
        col_letter = get_column_letter(col_num)
        sheet[col_letter + '2'] = header
    ###########################################
    # Merge cells for L52-Hard
    sheet.merge_cells('Z1:AC1')
    sheet['Z1'] = 'L52-Hard'
    sheet['Z1'].alignment = Alignment(horizontal='center')

    # Set column headers and values for L52-Hard
    for col_num, header in enumerate(headers, start=26):
        col_letter = get_column_letter(col_num)
        sheet[col_letter + '2'] = header
    #############################################
    # Merge cells for L52-Clay
    sheet.merge_cells('AD1:AG1')
    sheet['AD1'] = 'L52-Clay'
    sheet['AD1'].alignment = Alignment(horizontal='center')

    # Set column headers and values for L52-Clay
    for col_num, header in enumerate(headers, start=30):
        col_letter = get_column_letter(col_num)
        sheet[col_letter + '2'] = header
    #############################################
    # Merge cells for L52-Grass
    sheet.merge_cells('AH1:AK1')
    sheet['AH1'] = 'L52-Grass'
    sheet['AH1'].alignment = Alignment(horizontal='center')

    # Set column headers and values for L52-Grass
    for col_num, header in enumerate(headers, start=34):
        col_letter = get_column_letter(col_num)
        sheet[col_letter + '2'] = header
    #############################################
    # Merge cells for Career-ITF
    sheet.merge_cells('AL1:AO1')
    sheet['AL1'] = 'Career-ITF'
    sheet['AL1'].alignment = Alignment(horizontal='center')

    # Set column headers and values for Career-ITF
    for col_num, header in enumerate(headers, start=38):
        col_letter = get_column_letter(col_num)
        sheet[col_letter + '2'] = header
    #############################################
    # Merge cells for ITF-2023
    sheet.merge_cells('AP1:AS1')
    sheet['AP1'] = 'ITF-2023'
    sheet['AP1'].alignment = Alignment(horizontal='center')

    # Set column headers and values for ITF-2023
    for col_num, header in enumerate(headers, start=42):
        col_letter = get_column_letter(col_num)
        sheet[col_letter + '2'] = header
    #############################################
    # Merge cells for ITF-2022
    sheet.merge_cells('AT1:AW1')
    sheet['AT1'] = 'ITF-2022'
    sheet['AT1'].alignment = Alignment(horizontal='center')

    # Set column headers and values for ITF-2022
    for col_num, header in enumerate(headers, start=46):
        col_letter = get_column_letter(col_num)
        sheet[col_letter + '2'] = header
    #############################################

    # Save the workbook
    workbook.save('./data/women.xlsx')
    return './data/women.xlsx'


