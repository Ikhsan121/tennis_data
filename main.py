from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from time import sleep
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from openpyxl import load_workbook
from scraping_men import player_men
from write_data_xlsx import write_spreadsheet
from openpyxl import Workbook
from scraping_women import player_women
import random

# prompts the user to choose male or female
column_data = []
while True:
    gender = input("Type M for men or W for women: ").upper()
    if gender == "M":
        # load the player name (men)
        # Specify the file path of the workbook
        file_path = './source/Players_List_Tennis.xlsx'

        # Load the workbook
        workbook = load_workbook(file_path)

        # Select the second sheet (index 1) or specify the sheet name
        sheet = workbook.worksheets[0]

        # Extract data from the second column of the second sheet and convert rows to a list
        column_data = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=2, max_col=2):
            for cell in row:
                column_data.append(cell.value)
        column_data.pop(0)
        break

    elif gender == "W":
        # load the player name (men)
        # Specify the file path of the workbook
        file_path = './source/Players_List_Tennis.xlsx'

        # Load the workbook
        workbook = load_workbook(file_path)

        # Select the second sheet (index 1) or specify the sheet name
        sheet = workbook.worksheets[1]

        # Extract data from the second column of the second sheet and convert rows to a list
        column_data = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=2, max_col=2):
            for cell in row:
                column_data.append(cell.value)
        column_data.pop(0)
        break
    else:
        print("Invalid input. Please enter a valid input.")


#################################################################################
# setting the webdriver for chrome
service = Service(executable_path="C:\Development\chromedriver.exe")  # Path for Chrome web driver
options = Options()
options.add_argument("--start-maximized")
# options.add_argument("--headless")  # remove this if you want to see how the webdriver run.
options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36")
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(service=service, options=options)
wait = WebDriverWait(driver, 10)  # set wait object
actions = ActionChains(driver)
failed_name = []
print("Start...")
for i in range(len(column_data)):
    driver.get("https://www.tennisabstract.com/")
    player_search = driver.find_element(By.ID, 'tags')
    if gender == "M":
        try:
            if "-" in column_data[i]:
                name = column_data[i].split('-')[0].replace("\xa0", " ") + " " + column_data[i].split('-')[1].replace("\xa0", " ")
            else:
                name = column_data[i].replace("\xa0", " ")
            player_search.send_keys(name)  # input the player name
            menu_item = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/ul/li/a')))
            actions.move_to_element(menu_item).double_click().perform()
            sleep(1)
            page_source = driver.page_source
            player_data = player_men(response=page_source, name=column_data[i])  # this is a list
            write_spreadsheet(player_data, row_num=i + 3, gender=gender)
        except TimeoutException:
            print(f'{column_data[i]}: failed')
            # create xlsx file for failed attempt
            failed_name.append(column_data[i])
            filename = f'./data/failed_name_{gender}.xlsx'
            workbook = Workbook()
            sheet = workbook.active
            for item in failed_name:
                sheet.append([item])
            workbook.save(filename)
    elif gender == "W":
        space_positions = []
        try:
            if "-" in column_data[i]:
                name = column_data[i].split('-')[0] + " " + column_data[i].split('-')[1]
            else:
                name = column_data[i].replace("\xa0", " ")
            player_search.send_keys(name)  # input the player name
            menu_item = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/ul/li/a')))
            actions.move_to_element(menu_item).double_click().perform()
            sleep(1)
            page_source = driver.page_source
            player_data = player_women(response=page_source, name=column_data[i])  # this is a list
            write_spreadsheet(player_data, row_num=i + 3, gender=gender)
        except TimeoutException:
            print(f'{column_data[i]}: failed')
            # create xlsx file for failed attempt
            failed_name.append(column_data[i])
            filename = f'./data/failed_name_{gender}.xlsx'
            workbook = Workbook()
            sheet = workbook.active
            for item in failed_name:
                sheet.append([item])
            workbook.save(filename)

print("Completed.")


